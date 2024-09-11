import locale
from django.shortcuts import render,redirect
from django.contrib.auth import authenticate, login as nassib_login, logout
from django.utils import timezone
from django.contrib.auth.models import Permission
from django.utils import timezone
import random
from django.contrib.auth.models import User, Group, Permission
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required, permission_required
from django.shortcuts import render,redirect, get_object_or_404
from django.db.models import Q, Sum
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import PermissionDenied
from django.http import Http404, JsonResponse
from datetime import date, datetime, timedelta
from openpyxl import load_workbook
from nassib.models import Account, Achat, Agency, BatchProduction, CashBank, ComptaFp, ComptaPos, ComptaVm, Consolidation, Distribution, FinishedProduction, OthersRenewed, Product, Production, Provider, SalesControl, StockAgency, StockFP, StockGlobal, Transfer, TrialBalance, VenteMarchandises
from decimal import Decimal, InvalidOperation


def parse_decimal(value):
    try:
        return Decimal(str(value).replace(',', '.'))
    except InvalidOperation:
        return Decimal(0)
    
def error(request, exception=None):
    if isinstance(exception, PermissionDenied):
        status_code = 403
    elif isinstance(exception, Http404):
        status_code = 404
    else:
        status_code = getattr(exception, 'status_code', 500)

    return render(request, 'error.html', {'status_code': status_code}, status=status_code)

def user_login(request):
    context = {}
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            user.last_login = timezone.now()
            user.save()
            nassib_login(request, user)
            if user.is_superuser:
                return redirect('view_user')
            else:
                user_permission_ids = user.user_permissions.values_list('id', flat=True)
                permissions = Permission.objects.filter(id__in=user_permission_ids, codename__startswith='view')
                group_permissions = Permission.objects.filter(group__user=user,codename__startswith='view').distinct()
                if permissions.exists():
                    perm = random.choice(permissions)
                    return redirect(perm.codename)
                elif group_permissions.exists():
                        perm = random.choice(group_permissions)
                        return redirect(perm.codename)
                else:
                    context['error_message'] = 'Aucune permission de vue trouvée.'
                    return render(request, 'login.html', context)
        else:
            context['error_message'] = 'Nom d\'utilisateur ou mot de passe incorrect'
            return render(request, 'login.html', context)
    return render(request, 'login.html',context)

@login_required
def user_logout(request):
    logout(request)
    return redirect('login')

@permission_required('auth.view_user', raise_exception=True)
def view_user(request):
    search_query = request.POST.get('search', '')
    users = User.objects.all()
    if search_query:
        users = users.filter(Q(first_name__icontains=search_query) | Q(last_name__icontains=search_query) | Q(username__icontains=search_query) | Q(email__icontains=search_query))
    users = users.order_by('-id')
    paginator = Paginator(users, 12)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'users': page_obj,
        'username':username_current,
        'search_query':search_query
    }
    return render(request, 'user/view_user.html', context)

@permission_required('auth.add_user', raise_exception=True)
def add_user(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
    }

    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        confirm_password = request.POST['confirm_password']

        if password != confirm_password:
            context['error_message'] = 'Les mots de passe ne correspondent pas.'
            return render(request, 'user/add_user.html', context)

        if User.objects.filter(username=username).exists():
            context['error_message'] = 'Le nom d\'utilisateur existe déjà.'
            return render(request, 'user/add_user.html', context)

        new_user = User.objects.create_user(username=username, password=password)
        new_user.is_active = True
        new_user.is_staff = True
        new_user.save()

        return redirect('change_user', new_user.id)

    return render(request, 'user/add_user.html', context)

@permission_required('auth.change_user', raise_exception=True)
def change_user(request, user_id):
    user = get_object_or_404(User, id=user_id)

    nassib_and_auth_content_types = ContentType.objects.filter(
        Q(app_label='nassib') | Q(app_label='auth', model='user')
    ).exclude(model__in=['logentry'])

    user_permissions = user.user_permissions.filter(content_type__in=nassib_and_auth_content_types)

    unassigned_permissions = Permission.objects.filter(content_type__in=nassib_and_auth_content_types).exclude(
    id__in=user_permissions.values_list('id', flat=True)
    ).exclude(
        Q(content_type__app_label='nassib', content_type__model='stockglobal', codename__in=['add_stockglobal', 'change_stockglobal','delete_stockglobal']) |
        Q(content_type__app_label='nassib', content_type__model='stockagency', codename__in=['add_stockagency', 'change_stockagency','delete_stockagency']) | 
        Q(content_type__app_label='nassib', content_type__model='dashboard', codename__in=['add_dashboard', 'change_dashboard','delete_dashboard'])| 
        Q(content_type__app_label='nassib', content_type__model='trialbalance', codename__in=['add_trialbalance', 'change_trialbalance','delete_trialbalance'])| 
        Q(content_type__app_label='nassib', content_type__model='consolidation', codename__in=['add_consolidation', 'change_consolidation','delete_consolidation'])
    )
    all_groups = Group.objects.all()
    user_groups = user.groups.all()
    unassigned_groups = all_groups.difference(user_groups)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username':username_current,
        'user_permissions': user_permissions,
        'groups': unassigned_groups,
        'user_groups': user_groups,
        'permissions': unassigned_permissions,
        'user_change': user,
    }
    if request.method == 'POST':
        username = request.POST['username']
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        email = request.POST['email']
        is_active = request.POST.get('is_active', False) == 'on'
        is_superuser = request.POST.get('is_superuser', False) == 'on'
        if user.username != username:
            if User.objects.filter(username=username).exists():
                context['error_message'] = 'Le nom d\'utilisateur existe déjà.'
                return render(request, 'user/change_user.html', context)
        
        if user.email != email:
            if User.objects.filter(email=email).exists():
                context['error_message'] = 'L\'adresse e-mail existe déjà.'
                return render(request, 'user/change_user.html', context)
        user.username = username
        user.first_name = first_name
        user.last_name = last_name
        user.email = email
        user.is_active = is_active
        user.is_superuser = is_superuser
        user.save()
        aut_rem_options = request.POST.getlist('aut_rem')
        for permission_id in aut_rem_options:
            permission = Permission.objects.get(id=permission_id)
            if permission in unassigned_permissions:
                user.user_permissions.add(permission)
        for permission in user_permissions:
            if str(permission.id) not in aut_rem_options:
                user.user_permissions.remove(permission)
        group_rem_options = request.POST.getlist('group_rem')
        for group_id in group_rem_options:
            group = Group.objects.get(id=group_id)
            if group in unassigned_groups:
                user.groups.add(group)
        for group in user_groups:
            if str(group.id) not in group_rem_options:
                user.groups.remove(group)
        return redirect('view_user')
    return render(request, "user/change_user.html", context)

@login_required
def change_password(request, user_id):
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {}
    user = get_object_or_404(User, id=user_id)
    context = {
        'user_change': user,
        'username':username_current,
        'user_id':user_id
    }
    if request.method == 'POST':
        password = request.POST['password']
        confirm_password = request.POST['confirm_password']
        if password != confirm_password:
            context['error_message'] = 'Les mots de passe ne correspondent pas.'
            return render(request, 'password.html', context)
        user.set_password(password)
        user.save()
        return redirect('change_user', user_id=user_id)
    return render(request, 'password.html', context)

@permission_required('auth.delete_user', raise_exception=True)
def delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete()
    return redirect('view_user')

@permission_required('auth.view_group', raise_exception=True)
def view_group(request):
    search_query = request.POST.get('search', '')
    groups = Group.objects.all()
    if search_query:  
        groups = groups.filter(name__icontains=search_query)

    groups = groups.order_by('-id')
    paginator = Paginator(groups, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'groups': page_obj,
        'search_query':search_query
    }
    return render(request, 'group/view_group.html', context)

@permission_required('auth.change_group', raise_exception=True)
def change_group(request, group_id):
    group = get_object_or_404(Group, id=group_id)

    nassib_and_auth_content_types = ContentType.objects.filter(
        Q(app_label='nassib') | Q(app_label='auth', model='user')
    ).exclude(model__in=['logentry'])

    group_permissions = group.permissions.filter(content_type__in=nassib_and_auth_content_types)

    unassigned_permissions = Permission.objects.filter(content_type__in=nassib_and_auth_content_types).exclude(
    id__in=group_permissions.values_list('id', flat=True)
    ).exclude(
        Q(content_type__app_label='nassib', content_type__model='stockglobal', codename__in=['add_stockglobal', 'change_stockglobal','delete_stockglobal']) |
        Q(content_type__app_label='nassib', content_type__model='stockagency', codename__in=['add_stockagency', 'change_stockagency','delete_stockagency']) | 
        Q(content_type__app_label='nassib', content_type__model='dashboard', codename__in=['add_dashboard', 'change_dashboard','delete_dashboard'])| 
        Q(content_type__app_label='nassib', content_type__model='trialbalance', codename__in=['add_trialbalance', 'change_trialbalance','delete_trialbalance'])| 
        Q(content_type__app_label='nassib', content_type__model='consolidation', codename__in=['add_consolidation', 'change_consolidation','delete_consolidation'])
    )
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'group': group, 
        'permissions': unassigned_permissions, 
        'group_permissions': group_permissions
    }
    if request.method == 'POST':
        name = request.POST['name']
        if group.name != name:
            if Group.objects.filter(name=name).exists():
                context['error_message'] = 'Le nom du groupe existe déjà.'
                return render(request, 'group/change_group.html', context)
        group.name = name
        group.save()
        aut_rem_options = request.POST.getlist('aut_rem')
        
        for permission_id in aut_rem_options:
            permission = Permission.objects.get(id=permission_id)
            if permission in unassigned_permissions:
                group.permissions.add(permission)

        for permission in group_permissions:
            if str(permission.id) not in aut_rem_options:
                group.permissions.remove(permission)
        return redirect('view_group')
    
    return render(request, 'group/change_group.html', context)

@permission_required('auth.delete_group', raise_exception=True)
def delete_group(request, group_id):
    group = get_object_or_404(Group, id=group_id)
    group.permissions.clear()
    group.delete()
    return redirect('view_group')

@permission_required('auth.add_group', raise_exception=True)
def add_group(request):
    amana_and_auth_content_types = ContentType.objects.filter(
        Q(app_label='nassib') | Q(app_label='auth', model='user')
    ).exclude(model__in=['logentry'])

    permissions = Permission.objects.filter(content_type__in=amana_and_auth_content_types).exclude(
        Q(content_type__app_label='nassib', content_type__model='stockglobal', codename__in=['add_stockglobal', 'change_stockglobal','delete_stockglobal']) |
        Q(content_type__app_label='nassib', content_type__model='stockagency', codename__in=['add_stockagency', 'change_stockagency','delete_stockagency']) | 
        Q(content_type__app_label='nassib', content_type__model='dashboard', codename__in=['add_dashboard', 'change_dashboard','delete_dashboard'])| 
        Q(content_type__app_label='nassib', content_type__model='trialbalance', codename__in=['add_trialbalance', 'change_trialbalance','delete_trialbalance'])| 
        Q(content_type__app_label='nassib', content_type__model='consolidation', codename__in=['add_consolidation', 'change_consolidation','delete_consolidation'])
    )
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'permissions': permissions,
    }
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            if Group.objects.filter(name=name).exists():
                context['error_message'] = 'Le nom du groupe existe déjà.'
                return render(request, 'group/add_group.html', context)
            new_group = Group.objects.create(name=name)
            aut_rem_options = request.POST.getlist('aut_rem')
            for permission_id in aut_rem_options:
                new_group.permissions.add(permission_id)
            return redirect('view_group')
    return render(request, 'group/add_group.html', context)

@permission_required('nassib.view_stockglobal', raise_exception=True)
def view_stockglobal(request):
    search_query = request.POST.get('search', '')
    stockglobals = StockGlobal.objects.all()
    if search_query:  
        stockglobals = stockglobals.filter(Q(ingredient__icontains=search_query))

    stockglobals = stockglobals.order_by('ingredient')
    paginator = Paginator(stockglobals, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'stockglobals': page_obj,
        'search_query':search_query
    }
    return render(request, 'stockglobal/view_stockglobal.html', context)

@permission_required('nassib.view_stockagency', raise_exception=True)
def view_stockagency(request):
    search_query = request.POST.get('search', '')
    selected_agency = request.POST.get('agency', '') 
    stockagencies = StockAgency.objects.all()
    agencies = StockAgency.objects.values('agency').distinct()
    if search_query:  
        stockagencies = stockagencies.filter(Q(ingredient__icontains=search_query))
    if selected_agency:
        stockagencies = stockagencies.filter(agency__icontains=selected_agency)
    stockagencies = stockagencies.order_by('agency')
    paginator = Paginator(stockagencies, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'stockagencies': page_obj,
        'search_query':search_query,
        'agencies':agencies,
        'selected_agency': selected_agency
    }
    return render(request, 'stockagency/view_stockagency.html', context)

@permission_required('nassib.view_achat', raise_exception=True)
def view_achat(request):
    search_query = request.POST.get('search', '')
    achats = Achat.objects.all()
    if search_query:  
        achats = achats.filter(Q(provider__icontains=search_query) | Q(reference__icontains=search_query) | Q(agency__icontains=search_query))

    achats = achats.order_by('-id')
    paginator = Paginator(achats, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'achats': page_obj,
        'search_query':search_query
    }
    return render(request, 'achat/view_achat.html', context)

@permission_required('nassib.add_achat', raise_exception=True)
def add_achat(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    agencies = Agency.objects.all().order_by('agency')
    providers = Provider.objects.all().order_by('name_provider')
    accounts = Account.objects.all().order_by('account')
    
    context = {
        'username': username_current,
        'agencies': agencies,
        'providers': providers,
        'accounts': accounts,
    }

    if request.method == 'POST':
        agency = request.POST['agency']
        department = request.POST['department']
        ingredient = request.POST['ingredient']
        reference = request.POST['reference']
        provider = request.POST['provider']
        quantity = parse_decimal(request.POST['quantity'])
        unit_price = parse_decimal(request.POST['unit_price'])
        amount = parse_decimal(request.POST['amount'])
        payment_method = request.POST['payment_method']
        journal = request.POST['journal']
        initiator = request.POST['initiator']
        beneficiary = request.POST['beneficiary']
        debit_account = request.POST['debit_account']
        debit_amount = parse_decimal(request.POST['debit_amount'])
        credit_account = request.POST['credit_account']
        credit_amount = parse_decimal(request.POST['credit_amount'])
        stock_in = parse_decimal(request.POST['stock_in'])
        
        if journal != "ACHAT":
            context['error_message'] = "le journal doit etre l\'option 'ACHAT'"
            return render(request, 'achat/add_achat.html', context)
        
        achat = Achat.objects.create(
            agency=agency,
            department=department,
            ingredient=ingredient,
            reference=reference,
            provider=provider,
            quantity=quantity,
            unit_price=unit_price,
            amount=amount,
            payment_method=payment_method,
            journal=journal,
            initiator=initiator,
            beneficiary=beneficiary,
            debit_account=debit_account,
            debit_amount=debit_amount,
            credit_account=credit_account,
            credit_amount=credit_amount,
            stock_in=stock_in,
            date=timezone.now()
        )
        achat.save()
        
        try:
            stockglobal = StockGlobal.objects.get(ingredient=ingredient)
            stockglobal.amount += amount
            stockglobal.quantity += quantity
            stockglobal.stock_in += stock_in
            stockglobal.avg_price = stockglobal.amount / stockglobal.quantity 
        except StockGlobal.DoesNotExist:
            avg_price = amount / quantity
            stockglobal = StockGlobal(
                ingredient=ingredient,
                amount=amount,
                quantity=quantity,
                stock_in=stock_in,
                avg_price=avg_price,
                stock_out=0
            )
        stockglobal.save()

        try:
            stockagency = StockAgency.objects.get(ingredient=ingredient, agency=beneficiary)
            stockagency.quantity += quantity
            stockagency.stock_in += stock_in
            stockagency.last_update_date = timezone.now()
        except StockAgency.DoesNotExist:
            stockagency = StockAgency(
                agency=achat.beneficiary,
                ingredient=ingredient,
                quantity=quantity,
                stock_in=stock_in,
                stock_out=0,
                last_update_date=timezone.now()
            )
        stockagency.save()

        return redirect('view_achat')

    return render(request, 'achat/add_achat.html', context)

    
@permission_required('nassib.change_achat', raise_exception=True)
def change_achat(request, achat_id):
    achat = get_object_or_404(Achat, id=achat_id)
    old_ingredient = achat.ingredient
    old_amount = parse_decimal(str(achat.amount))
    old_quantity = parse_decimal(str(achat.quantity))
    old_stock_in = parse_decimal(str(achat.stock_in))
    old_beneficiary = achat.beneficiary
    old_agency = achat.agency
    old_initiator = achat.initiator
    current_user = request.user
    username_current = current_user.username.capitalize()
    agencies = Agency.objects.all().order_by('agency')
    providers = Provider.objects.all().order_by('name_provider')
    accounts = Account.objects.all().order_by('account')
    
    context = {
        'username': username_current,
        'agencies': agencies,
        'providers': providers,
        'accounts': accounts,
        'achat': achat,
    }

    if request.method == 'POST':
        agency = request.POST['agency']
        department = request.POST['department']
        ingredient = request.POST['ingredient']
        reference = request.POST['reference']
        provider = request.POST['provider']
        quantity = parse_decimal(request.POST['quantity'])
        unit_price = parse_decimal(request.POST['unit_price'])
        amount = parse_decimal(request.POST['amount'])
        payment_method = request.POST['payment_method']
        journal = request.POST['journal']
        initiator = request.POST['initiator']
        beneficiary = request.POST['beneficiary']
        debit_account = request.POST['debit_account']
        debit_amount = parse_decimal(request.POST['debit_amount'])
        credit_account = request.POST['credit_account']
        credit_amount = parse_decimal(request.POST['credit_amount'])
        stock_in = parse_decimal(request.POST['stock_in'])
        
        if journal != "ACHAT":
            context['error_message'] = "le journal doit etre l\'option 'ACHAT'"
            return render(request, 'achat/change_achat.html', context)
        
        achat.agency = agency
        achat.department = department
        achat.ingredient = ingredient
        achat.reference = reference
        achat.provider = provider
        achat.quantity = quantity
        achat.unit_price = unit_price
        achat.amount = amount
        achat.payment_method = payment_method
        achat.journal = journal
        achat.initiator = initiator
        achat.beneficiary = beneficiary
        achat.debit_account = debit_account
        achat.debit_amount = debit_amount
        achat.credit_account = credit_account
        achat.credit_amount = credit_amount
        achat.stock_in = stock_in
        achat.date = timezone.now()
        achat.save()
        
        if old_ingredient == ingredient and old_beneficiary == beneficiary and old_agency == agency and old_initiator == initiator:
            stockglobal = StockGlobal.objects.get(ingredient=ingredient)
            stockglobal.amount += amount - old_amount
            stockglobal.quantity += quantity - old_quantity
            stockglobal.stock_in += stock_in - old_stock_in
            
            if stockglobal.quantity != 0:
                stockglobal.avg_price = stockglobal.amount / stockglobal.quantity
            else:
                stockglobal.avg_price = Decimal('0.0')
                
            stockglobal.save()

            stockagency = StockAgency.objects.get(ingredient=ingredient, agency=beneficiary)
            stockagency.quantity += quantity - old_quantity
            stockagency.stock_in += stock_in - old_stock_in
            stockagency.last_update_date = timezone.now()
            stockagency.save()
        else:
            old_stockglobal = StockGlobal.objects.get(ingredient=old_ingredient)
            old_stockglobal.amount -= old_amount
            old_stockglobal.quantity -= old_quantity
            old_stockglobal.stock_in -= old_stock_in
            if old_stockglobal.quantity != parse_decimal('0'):
                old_stockglobal.avg_price = old_stockglobal.amount / old_stockglobal.quantity
                old_stockglobal.save()
            else:
                old_stockglobal.delete()
                
            stockglobal = StockGlobal.objects.filter(ingredient=ingredient)
            if stockglobal.exists():
                stockglobal.amount += amount
                stockglobal.quantity += quantity
                stockglobal.stock_in += stock_in
                stockglobal.avg_price = stockglobal.amount / stockglobal.quantity 
            else:
                avg_price = amount / quantity
                stockglobal = StockGlobal(
                    ingredient=ingredient,
                    amount=amount,
                    quantity=quantity,
                    stock_in=stock_in,
                    avg_price=avg_price,
                    stock_out=0
                )
            stockglobal.save()
            
            old_stockagency = StockAgency.objects.get(ingredient=old_ingredient, agency=old_beneficiary)
            old_stockagency.quantity -= old_quantity
            old_stockagency.stock_in -= old_stock_in
            old_stockagency.last_update_date = timezone.now()
            if old_stockagency.quantity != parse_decimal('0'):
                old_stockagency.save()
            else:
                old_stockagency.delete()
                
            stockagency = StockAgency.objects.filter(ingredient=ingredient, agency=beneficiary)
            if stockagency.exists():
                stockagency.quantity += quantity
                stockagency.stock_in += stock_in
                stockagency.last_update_date = timezone.now()
            else:
                stockagency = StockAgency(
                    agency=achat.beneficiary,
                    ingredient=ingredient,
                    quantity=quantity,
                    stock_in=stock_in,
                    stock_out=0,
                    last_update_date=timezone.now()
                )
            stockagency.save()

        return redirect('view_achat')

    return render(request, 'achat/change_achat.html', context)

@permission_required('nassib.add_achat', raise_exception=True)
def import_achat(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
    }

    if request.method == 'POST':
        excel_file = request.POST['excel_file']
        try:
            workbook = load_workbook(excel_file)
            sheet = workbook.active
            
            rows = iter(sheet.rows)
            next(rows) 
            
            for row in rows:
                reference = row[0].value
                agency = row[1].value
                department = row[3].value
                ingredient = row[4].value
                provider = row[5].value
                quantity = parse_decimal(row[6].value)
                unit_price = parse_decimal(row[7].value)
                amount = parse_decimal(row[8].value)
                payment_method = row[9].value
                journal = row[10].value
                debit_account = row[11].value
                debit_amount = parse_decimal(row[12].value)
                initiator = row[13].value
                credit_account = row[14].value
                credit_amount = parse_decimal(row[15].value)
                beneficiary = row[16].value
                stock_in = parse_decimal(row[17].value)
                date = row[18].value
                
                achat = Achat.objects.create(
                    agency=agency,
                    department=department,
                    ingredient=ingredient,
                    reference=reference,
                    provider=provider,
                    quantity=quantity,
                    unit_price=unit_price,
                    amount=amount,
                    payment_method=payment_method,
                    journal=journal,
                    initiator=initiator,
                    beneficiary=beneficiary,
                    debit_account=debit_account,
                    debit_amount=debit_amount,
                    credit_account=credit_account,
                    credit_amount=credit_amount,
                    stock_in=stock_in,
                    date=date
                )
                achat.save()
                
                try:
                    stockglobal = StockGlobal.objects.get(ingredient=ingredient)
                    stockglobal.amount += amount
                    stockglobal.quantity += quantity
                    stockglobal.stock_in += stock_in
                    stockglobal.avg_price = stockglobal.amount / stockglobal.quantity 
                except StockGlobal.DoesNotExist:
                    avg_price = amount / quantity
                    stockglobal = StockGlobal(
                        ingredient=ingredient,
                        amount=amount,
                        quantity=quantity,
                        stock_in=stock_in,
                        avg_price=avg_price,
                        stock_out=0
                    )
                stockglobal.save()

                try:
                    stockagency = StockAgency.objects.get(ingredient=ingredient, agency=beneficiary)
                    stockagency.quantity += quantity
                    stockagency.stock_in += stock_in
                    stockagency.last_update_date = timezone.now()
                except StockAgency.DoesNotExist:
                    stockagency = StockAgency(
                        agency=achat.beneficiary,
                        ingredient=ingredient,
                        quantity=quantity,
                        stock_in=stock_in,
                        stock_out=0,
                        last_update_date=timezone.now()
                    )
                stockagency.save()
            return redirect('view_achat')
        except Exception as e:
            context['error_message'] = f'Une erreur s\'est produite lors de l\'importation des données : {str(e)}'
            return render(request, 'achat/import_achat.html', context)

    return render(request, 'achat/import_achat.html', context)

@permission_required('nassib.delete_achat', raise_exception=True)
def delete_achat(request, achat_id):
    achat = get_object_or_404(Achat, id=achat_id)
    
    if achat.date == date.today():
        stockglobal = StockGlobal.objects.get(ingredient=achat.ingredient)
        stockglobal.amount -= achat.amount
        stockglobal.quantity -= achat.quantity
        stockglobal.stock_in -= achat.stock_in

        if stockglobal.quantity != 0:
            stockglobal.avg_price = stockglobal.amount / stockglobal.quantity
        else:
            stockglobal.avg_price = parse_decimal(0)
        
        stockglobal.save()
        
        stockagency = StockAgency.objects.get(ingredient=achat.ingredient, agency=achat.beneficiary)
        stockagency.quantity -= achat.quantity
        stockagency.stock_in -= achat.stock_in
        stockagency.last_update_date = timezone.now()
        stockagency.save()

        if stockglobal.stock_in == parse_decimal(0):
            stockglobal.delete()
        
        if stockagency.stock_in == parse_decimal(0):
            stockagency.delete()
        
        achat.delete()
    else:
        achat.delete()
    
    return redirect('view_achat')

@login_required
def get_account(request, account):
    Label = ''

    try:
        account = get_object_or_404(Account, account=int(account))
        Label = account.Label
    except Account.DoesNotExist:
        pass
    except ValueError:
        pass

    return JsonResponse({'label': Label})

@login_required
def search_ingredient(request,ingredient):
    ingredients = StockGlobal.objects.filter(ingredient__icontains=ingredient)
    ingredient_list = list(ingredients.values('ingredient'))
    return JsonResponse(ingredient_list, safe=False)

@permission_required('nassib.view_transfer', raise_exception=True)
def view_transfer(request):
    search_query = request.POST.get('search', '')
    transfers = Transfer.objects.all()
    if search_query:  
        transfers = transfers.filter(Q(no_bl__icontains=search_query) | Q(initiator__icontains=search_query) | Q(beneficiary__icontains=search_query))

    transfers = transfers.order_by('-id')
    paginator = Paginator(transfers, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'transfers': page_obj,
        'search_query':search_query
    }
    return render(request, 'transfer/view_transfer.html', context)

@permission_required('nassib.add_transfer', raise_exception=True)
def add_transfer(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    agencies = Agency.objects.all().order_by('agency')
    
    context = {
        'username': username_current,
        'agencies': agencies,
    }

    if request.method == 'POST':
        date = request.POST['date']
        no_bl = request.POST['no_bl']
        ingredient = request.POST['ingredient']
        quantity_issued = parse_decimal(request.POST['quantity_issued'])
        amount_issue = parse_decimal(request.POST['amount_issue'])
        journal = request.POST['journal']
        initiator = request.POST['initiator']
        beneficiary = request.POST['beneficiary']
        status = request.POST['status']
        stock_out = parse_decimal(request.POST['stock_out'])
        
        if status == "Confirmez":
            context['error_message'] = "le statut doit etre l\'option 'attente'"
            return render(request, 'transfer/add_transfer.html', context)
        
        if journal != "TRANSFERT":
            context['error_message'] = "le journal doit etre l\'option 'TRANSFERT'"
            return render(request, 'transfer/add_transfer.html', context)
        
        transfer = Transfer.objects.create(
            date=date,
            no_bl=no_bl,
            ingredient=ingredient,
            quantity_issued=quantity_issued,
            amount_issue=amount_issue,
            status=status,
            stock_out=stock_out,
            journal=journal,
            initiator=initiator,
            beneficiary=beneficiary,
            quantity_recu = 0,
            stock_in = stock_out,
            quantity_ecart = stock_out
        )
        transfer.save()
        
        try:
            stockagency_initiator = StockAgency.objects.get(ingredient=ingredient, agency=initiator)
            if quantity_issued > stockagency_initiator.quantity:
                context['error_message'] = "la quantité emis est plus grande que la quantité en stock pour cette agence."
                return render(request, 'transfer/add_transfer.html', context)
            stockagency_initiator.stock_out += stock_out
            stockagency_initiator.quantity -= quantity_issued
            stockagency_initiator.last_update_date = timezone.now()
        except StockAgency.DoesNotExist:
            context['error_message'] = "Une erreur s\'est produite lors de la recherche des données de cette agence."
            return render(request, 'transfer/add_transfer.html', context)
        stockagency_initiator.save()
        
        return redirect('view_transfer')

    return render(request, 'transfer/add_transfer.html', context)

@permission_required('nassib.delete_transfer', raise_exception=True)
def delete_transfer(request,transfer_id):
    transfer = get_object_or_404(Transfer, id=transfer_id)
    status = transfer.status
    if status == "Confirmez":
        if transfer.date == date.today():
            
            stockagency_beneficiary = StockAgency.objects.get(ingredient=transfer.ingredient, agency=transfer.beneficiary)
            stockagency_beneficiary.stock_in -= transfer.stock_in
            stockagency_beneficiary.quantity = stockagency_beneficiary.stock_in - stockagency_beneficiary.stock_out
            stockagency_beneficiary.last_update_date = timezone.now()
            stockagency_beneficiary.save()
            
            if stockagency_beneficiary.stock_in == parse_decimal(0):
                stockagency_beneficiary.delete()
            
            stockagency_initiator = StockAgency.objects.get(ingredient=transfer.ingredient, agency=transfer.initiator)
            stockagency_initiator.stock_out -= transfer.stock_in
            stockagency_initiator.quantity = stockagency_initiator.stock_in - stockagency_initiator.stock_out
            stockagency_initiator.last_update_date = timezone.now()
            stockagency_initiator.save()
            
            transfer.delete()
        else:
            transfer.delete()
    else:
        stockagency_initiator = StockAgency.objects.get(ingredient=transfer.ingredient, agency=transfer.initiator)
        stockagency_initiator.stock_out -= transfer.stock_in
        stockagency_initiator.quantity = stockagency_initiator.stock_in - stockagency_initiator.stock_out
        stockagency_initiator.last_update_date = timezone.now()
        stockagency_initiator.save()
        transfer.delete()
    return redirect('view_transfer')

@login_required
def get_ingredients_by_initiator(request, initiator):
    ingredients = StockAgency.objects.filter(agency=initiator, quantity__gt=0).values_list('ingredient', flat=True).distinct()
    return JsonResponse(list(ingredients), safe=False)

@login_required
def get_ingredient_avgprice(request, ingredient):
    try:
        stock_global = StockGlobal.objects.get(ingredient=ingredient)
        avg_price = stock_global.avg_price
    except StockGlobal.DoesNotExist:
        avg_price = 0
    return JsonResponse({'avg_price': avg_price})

@permission_required('nassib.view_transfer', raise_exception=True)
def view_validation(request):
    search_query = request.POST.get('search', '')
    transfers = Transfer.objects.filter(status="En attente")
    if search_query:  
        transfers = transfers.filter(Q(no_bl__icontains=search_query) | Q(initiator__icontains=search_query) | Q(beneficiary__icontains=search_query))

    transfers = transfers.order_by('-id')
    paginator = Paginator(transfers, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'transfers': page_obj,
        'search_query':search_query
    }
    return render(request, 'validation/view_validation.html', context)

@permission_required('nassib.view_transfer', raise_exception=True)
def validation(request,transfer_id):
    transfer = get_object_or_404(Transfer, id=transfer_id)
    current_user = request.user
    username_current = current_user.username.capitalize()
    agencies = Agency.objects.all().order_by('agency')
    context = {
        'username': username_current,
        'transfer': transfer,
        'agencies':agencies
    }
    if request.method == 'POST':
        quantity_recu = parse_decimal(request.POST['quantity_recu'])
        status = request.POST['status']
        if status != "Confirmez":
            context['error_message'] = "le statut doit etre l\'option 'Confirmez'"
            return render(request, 'validation/validation.html', context)
        try:
            stockagency_beneficiary = StockAgency.objects.get(ingredient=transfer.ingredient, agency=transfer.beneficiary)
            stockagency_beneficiary.quantity += quantity_recu
            stockagency_beneficiary.stock_in += quantity_recu
            stockagency_beneficiary.last_update_date = timezone.now()
        except StockAgency.DoesNotExist:
            stockagency_beneficiary = StockAgency(
                agency=transfer.beneficiary,
                ingredient=transfer.ingredient,
                quantity=quantity_recu,
                stock_in=quantity_recu,
                stock_out=0,
                last_update_date=timezone.now()
            )
        stockagency_beneficiary.save()
        transfer.quantity_recu = quantity_recu
        quantity_ecart = quantity_recu
        transfer.quantity_ecart = transfer.stock_in - quantity_ecart
        transfer.date = timezone.now()
        transfer.status = status
        transfer.save()
        return redirect('view_validation')
    return render(request, 'validation/validation.html', context)

@permission_required('nassib.view_ventemarchandises', raise_exception=True)
def view_ventemarchandises(request):
    search_query = request.POST.get('search', '')
    ventemarchandises = VenteMarchandises.objects.all()
    if search_query:  
        ventemarchandises = ventemarchandises.filter(Q(customer__icontains=search_query))

    ventemarchandises = ventemarchandises.order_by('-id')
    paginator = Paginator(ventemarchandises, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'ventemarchandises': page_obj,
        'search_query':search_query
    }
    return render(request, 'ventemarchandises/view_ventemarchandises.html', context)

@permission_required('nassib.add_ventemarchandises', raise_exception=True)
def add_ventemarchandises(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    agencies = Agency.objects.all().order_by('agency')
    context = {
        'username': username_current,
        'agencies':agencies
    }
    if request.method == 'POST':
        date = request.POST['date']
        agency = request.POST['agency']
        customer = request.POST['customer']
        quantity  = parse_decimal(request.POST['quantity'])
        ingredient = request.POST['ingredient']
        cost_of_sales = parse_decimal(request.POST['cost_of_sales'])
        selling_price = parse_decimal(request.POST['selling_price'])
        stock_out = parse_decimal(request.POST['stock_out'])
        vante_marchandises = parse_decimal(request.POST['vante_marchandises'])
        marge_brute = parse_decimal(request.POST['marge_brute'])
        try:
            stockagency = StockAgency.objects.get(agency=agency, ingredient=ingredient)
            if stockagency.quantity < quantity:
                context['error_message'] = "La quantité est plus grand que la quantité en stock pour cette agence."
                return render(request, 'ventemarchandises/add_ventemarchandises.html', context)
            ventemarchandises = VenteMarchandises.objects.create(
                date=date,
                agency=agency,
                customer=customer,
                ingredient=ingredient,
                quantity =quantity,
                cost_of_sales=cost_of_sales,
                selling_price=selling_price,
                stock_out=stock_out,
                vante_marchandises=vante_marchandises,
                marge_brute=marge_brute
            )
            ventemarchandises.save()
            
            stockagency.stock_out += stock_out
            stockagency.quantity = stockagency.stock_in - stockagency.stock_out
            stockagency.last_update_date = timezone.now()
            stockagency.save()
            try:
                stockglobal = StockGlobal.objects.get(ingredient=ingredient)
                stockglobal.stock_out += stock_out
                stockglobal.quantity = stockglobal.stock_in - stockglobal.stock_out
                stockglobal.amount = stockglobal.quantity * stockglobal.avg_price
                stockglobal.save()
            except StockGlobal.DoesNotExist:
                context['error_message'] = "Une erreur s\'est produite lors de la recherche des données de cette ingredient."
                return render(request, 'ventemarchandises/add_ventemarchandises.html', context)
        except StockAgency.DoesNotExist:
            context['error_message'] = "Cette agence ne possede pas de stock pour cette produit selectionnez."
            return render(request, 'ventemarchandises/add_ventemarchandises.html', context)
        
        return redirect('view_ventemarchandises')
    return render(request, 'ventemarchandises/add_ventemarchandises.html', context)

@permission_required('nassib.change_ventemarchandises', raise_exception=True)
def change_ventemarchandises(request,vente_id):
    ventemarchandise = get_object_or_404(VenteMarchandises, id=vente_id)
    old_agency = ventemarchandise.agency
    old_ingredient = ventemarchandise.ingredient
    old_stock_out = parse_decimal(ventemarchandise.stock_out)
    current_user = request.user
    username_current = current_user.username.capitalize()
    agencies = Agency.objects.all().order_by('agency')
    context = {
        'username': username_current,
        'agencies':agencies,
        'ventemarchandise':ventemarchandise
    }
    if request.method == 'POST':
        date = request.POST['date']
        agency = request.POST['agency']
        customer = request.POST['customer']
        quantity  = parse_decimal(request.POST['quantity'])
        ingredient = request.POST['ingredient']
        cost_of_sales = parse_decimal(request.POST['cost_of_sales'])
        selling_price = parse_decimal(request.POST['selling_price'])
        stock_out = parse_decimal(request.POST['stock_out'])
        vante_marchandises = parse_decimal(request.POST['vante_marchandises'])
        marge_brute = parse_decimal(request.POST['marge_brute'])
        if old_agency == agency and old_ingredient == ingredient:
            try:
                old_stockagency = StockAgency.objects.get(agency=old_agency, ingredient=old_ingredient)
                if old_stockagency.quantity < quantity:
                    context['error_message'] = "La quantité est plus grand que la quantité en stock pour cette agence."
                    return render(request, 'ventemarchandises/change_ventemarchandises.html', context)
                
                ventemarchandise.date=date
                ventemarchandise.agency=agency
                ventemarchandise.customer=customer
                ventemarchandise.ingredient=ingredient
                ventemarchandise.quantity =quantity
                ventemarchandise.cost_of_sales=cost_of_sales
                ventemarchandise.selling_price=selling_price
                ventemarchandise.stock_out=stock_out
                ventemarchandise.vante_marchandises=vante_marchandises
                ventemarchandise.marge_brute=marge_brute
                ventemarchandise.save()
                
                old_stockagency.stock_out += stock_out - old_stock_out
                old_stockagency.quantity = old_stockagency.stock_in - old_stockagency.stock_out
                old_stockagency.last_update_date = timezone.now()
                old_stockagency.save()
                try:
                    stockglobal = StockGlobal.objects.get(ingredient=ingredient)
                    stockglobal.stock_out += stock_out - old_stock_out
                    stockglobal.quantity = stockglobal.stock_in - stockglobal.stock_out
                    stockglobal.amount = stockglobal.quantity * stockglobal.avg_price
                    stockglobal.save()
                except StockGlobal.DoesNotExist:
                    context['error_message'] = "Une erreur s\'est produite lors de la recherche des données de cette ingredient."
                    return render(request, 'ventemarchandises/change_ventemarchandises.html', context)
            except StockAgency.DoesNotExist:
                context['error_message'] = "Cette agence ne possede pas de stock pour cette produit selectionnez."
                return render(request, 'ventemarchandises/change_ventemarchandises.html', context)
        else:
            try:
                stockagency = StockAgency.objects.get(agency=agency, ingredient=ingredient)
                if stockagency.quantity < quantity:
                    context['error_message'] = "La quantité est plus grand que la quantité en stock pour cette agence."
                    return render(request, 'ventemarchandises/change_ventemarchandises.html', context)
                
                ventemarchandise.date=date
                ventemarchandise.agency=agency
                ventemarchandise.customer=customer
                ventemarchandise.ingredient=ingredient
                ventemarchandise.quantity =quantity
                ventemarchandise.cost_of_sales=cost_of_sales
                ventemarchandise.selling_price=selling_price
                ventemarchandise.stock_out=stock_out
                ventemarchandise.vante_marchandises=vante_marchandises
                ventemarchandise.marge_brute=marge_brute
                ventemarchandise.save()
                
                old_stockagency = StockAgency.objects.get(agency=old_agency, ingredient=old_ingredient)
                old_stockagency.stock_out -= old_stock_out
                old_stockagency.quantity = old_stockagency.stock_in - old_stockagency.stock_out
                old_stockagency.last_update_date = timezone.now()
                old_stockagency.save()
                
                stockagency.stock_out += stock_out
                stockagency.quantity = stockagency.stock_in - stockagency.stock_out
                stockagency.last_update_date = timezone.now()
                stockagency.save()
                try:
                    stockglobal = StockGlobal.objects.get(ingredient=ingredient)
                    stockglobal.stock_out += stock_out - old_stock_out
                    stockglobal.quantity = stockglobal.stock_in - stockglobal.stock_out
                    stockglobal.amount = stockglobal.quantity * stockglobal.avg_price
                    stockglobal.save()
                except StockGlobal.DoesNotExist:
                    context['error_message'] = "Une erreur s\'est produite lors de la recherche des données de cette ingredient."
                    return render(request, 'ventemarchandises/change_ventemarchandises.html', context)
            except StockAgency.DoesNotExist:
                context['error_message'] = "Cette agence ne possede pas de stock pour cette produit selectionnez."
                return render(request, 'ventemarchandises/change_ventemarchandises.html', context)
            
        return redirect('view_ventemarchandises')
    return render(request, 'ventemarchandises/change_ventemarchandises.html', context)

@permission_required('nassib.delete_ventemarchandises', raise_exception=True)
def delete_ventemarchandises(request, vente_id):
    ventemarchandise = get_object_or_404(VenteMarchandises, id=vente_id)
    
    if ventemarchandise.date == date.today():
        stockglobal = StockGlobal.objects.get(ingredient=ventemarchandise.ingredient)
        stockglobal.stock_out -= ventemarchandise.stock_out
        stockglobal.quantity = stockglobal.stock_in - stockglobal.stock_out
        stockglobal.amount = stockglobal.quantity * stockglobal.avg_price
        stockglobal.save()
        
        stockagency = StockAgency.objects.get(ingredient=ventemarchandise.ingredient, agency=ventemarchandise.agency)
        stockagency.stock_out -= ventemarchandise.stock_out
        stockagency.quantity = stockagency.stock_in - stockagency.stock_out
        stockagency.last_update_date = timezone.now()
        stockagency.save()
        
        ventemarchandise.delete()
    else:
        ventemarchandise.delete()
    
    return redirect('view_ventemarchandises')

@permission_required('nassib.view_provider', raise_exception=True)
def view_provider(request):
    search_query = request.POST.get('search', '')
    provider = Provider.objects.all()
    if search_query:  
        provider = provider.filter(Q(name_provider__icontains=search_query))

    provider = provider.order_by('name_provider')
    paginator = Paginator(provider, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'providers': page_obj,
        'search_query':search_query
    }
    return render(request, 'provider/view_provider.html', context)

@permission_required('nassib.add_provider', raise_exception=True)
def add_provider(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current
    }
    if request.method == 'POST':
        name_provider = request.POST['name_provider']
        provider = Provider.objects.filter(name_provider=name_provider)
        if provider.exists():
            context['error_message'] = 'Ce Fournisseur existe deja.'
            return render(request, 'provider/add_provider.html', context)
        
        provider = Provider.objects.create(
            name_provider=name_provider
        )
        provider.save()
        return redirect('view_provider')
    return render(request, 'provider/add_provider.html', context)

@permission_required('nassib.add_provider', raise_exception=True)
def import_provider(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current
    }

    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            context['error_message'] = 'Veuillez sélectionner un fichier à importer.'
            return render(request, 'provider/import_provider.html', context)
        
        excel_file = request.FILES['excel_file']
        try:
            workbook = load_workbook(excel_file)
            sheet = workbook.active
            
            rows = iter(sheet.rows)
            next(rows)  # Skip the header row
            for row in rows:
                name_provider = row[0].value
                p = Provider.objects.filter(name_provider=name_provider)
                if p.exists():
                    context['error_message'] = f'Le Fournisseur {name_provider} existe déjà.'
                    return render(request, 'provider/import_provider.html', context)
                provider = Provider.objects.create(name_provider=name_provider)
                provider.save()
            return redirect('view_provider')
        except Exception as e:
            context['error_message'] = f'Une erreur s\'est produite lors de l\'importation des données : {str(e)}'
            return render(request, 'provider/import_provider.html', context)
        
    return render(request, 'provider/import_provider.html', context)

@permission_required('nassib.change_provider', raise_exception=True)
def change_provider(request,provider_id):
    provider = get_object_or_404(Provider, id=provider_id)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'provider':provider
    }
    if request.method == 'POST':
        name_provider = request.POST['name_provider']
        if provider.name_provider != name_provider:
            p = Provider.objects.filter(name_provider=name_provider)
            if p.exists():
                context['error_message'] = 'Ce Fournisseur existe deja.'
                return render(request, 'provider/change_provider.html', context)
        provider.name_provider = name_provider
        provider.save()
        
        return redirect('view_provider')
    return render(request, 'provider/change_provider.html', context)

@permission_required('nassib.delete_provider', raise_exception=True)
def delete_provider(request, provider_id):
    provider = get_object_or_404(Provider, id=provider_id)
    provider.delete()
    return redirect('view_provider')

@permission_required('nassib.view_product', raise_exception=True)
def view_product(request):
    search_query = request.POST.get('search', '')
    product = Product.objects.all()
    if search_query:  
        product = product.filter(Q(item__icontains=search_query))

    product = product.order_by('item')
    paginator = Paginator(product, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'products': page_obj,
        'search_query':search_query
    }
    return render(request, 'product/view_product.html', context)

@permission_required('nassib.add_product', raise_exception=True)
def add_product(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current
    }
    if request.method == 'POST':
        item = request.POST['item']
        price = parse_decimal(request.POST['price'])
        product = Product.objects.filter(item=item)
        if product.exists():
            context['error_message'] = 'Ce Produit existe deja.'
            return render(request, 'provider/add_provider.html', context)
        
        product = Product.objects.create(
            item=item,
            price=price
        )
        product.save()
        return redirect('view_product')
    return render(request, 'product/add_product.html', context)

@permission_required('nassib.add_product', raise_exception=True)
def import_product(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current
    }

    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            context['error_message'] = 'Veuillez sélectionner un fichier à importer.'
            return render(request, 'product/import_product.html', context)
        
        excel_file = request.FILES['excel_file']
        try:
            workbook = load_workbook(excel_file)
            sheet = workbook.active
            
            rows = iter(sheet.rows)
            next(rows)  # Skip the header row
            for row in rows:
                item = row[0].value
                price = parse_decimal(row[1].value)
                p = Product.objects.filter(item=item)
                if p.exists():
                    context['error_message'] = f'Le Produit {item} existe déjà.'
                    return render(request, 'product/import_product.html', context)
                product = Product.objects.create(
                    item=item,
                    price=price
                )
                product.save()
            return redirect('view_product')
        except Exception as e:
            context['error_message'] = f'Une erreur s\'est produite lors de l\'importation des données : {str(e)}'
            return render(request, 'product/import_product.html', context)
        
    return render(request, 'product/import_product.html', context)

@permission_required('nassib.change_product', raise_exception=True)
def change_product(request,product_id):
    product = get_object_or_404(Product, id=product_id)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'product':product
    }
    if request.method == 'POST':
        item = request.POST['item']
        price = parse_decimal(request.POST['price'])
        if product.item != item:
            p = Product.objects.filter(item=item)
            if p.exists():
                context['error_message'] = 'Ce Produit existe deja.'
                return render(request, 'product/change_product.html', context)
        product.item = item
        product.price = price
        product.save()
        
        return redirect('view_product')
    return render(request, 'product/change_product.html', context)

@permission_required('nassib.delete_product', raise_exception=True)
def delete_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    product.delete()
    return redirect('view_product')

@permission_required('nassib.view_account', raise_exception=True)
def view_account(request):
    search_query = request.POST.get('search', '')
    account = Account.objects.all()
    if search_query:  
        account = account.filter(Q(account__icontains=search_query) | Q(family__icontains=search_query) | Q(Label__icontains=search_query))

    account = account.order_by('account')
    paginator = Paginator(account, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'accounts': page_obj,
        'search_query':search_query
    }
    return render(request, 'account/view_account.html', context)

@permission_required('nassib.add_account', raise_exception=True)
def add_account(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current
    }
    if request.method == 'POST':
        code = request.POST['account']
        family = request.POST['family']
        classe = request.POST['classe']
        Label = request.POST['Label']
        OHADA_Code = request.POST['OHADA_Code']
        note = request.POST['note']
        account = Account.objects.filter(account=code)
        if account.exists():
            context['error_message'] = 'Ce Compte OHADA existe deja.'
            return render(request, 'account/add_account.html', context)
        
        account = Account.objects.create(
            account=code,
            family=family,
            classe=classe,
            Label=Label,
            OHADA_Code=OHADA_Code,
            note=note
        )
        account.save()
        return redirect('view_account')
    return render(request, 'account/add_account.html', context)

@permission_required('nassib.add_account', raise_exception=True)
def import_account(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current
    }

    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            context['error_message'] = 'Veuillez sélectionner un fichier à importer.'
            return render(request, 'account/import_account.html', context)
        
        excel_file = request.FILES['excel_file']
        try:
            workbook = load_workbook(excel_file)
            sheet = workbook.active
            
            rows = iter(sheet.rows)
            next(rows)  # Skip the header row
            for row in rows:
                code = row[0].value
                family = row[1].value
                classe = row[2].value
                Label = row[3].value
                OHADA_Code = row[4].value
                note = row[5].value
                a = Account.objects.filter(account=code)
                if a.exists():
                    context['error_message'] = f'Le Compte OHADA {code} existe déjà.'
                    return render(request, 'account/import_account.html', context)
                account = Account.objects.create(
                    account=code,
                    family=family,
                    classe=classe,
                    Label=Label,
                    OHADA_Code=OHADA_Code,
                    note=note
                )
                account.save()
            return redirect('view_account')
        except Exception as e:
            context['error_message'] = f'Une erreur s\'est produite lors de l\'importation des données : {str(e)}'
            return render(request, 'account/import_account.html', context)
        
    return render(request, 'account/import_account.html', context)

@permission_required('nassib.change_account', raise_exception=True)
def change_account(request,account_id):
    account = get_object_or_404(Account, id=account_id)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'account':account
    }
    if request.method == 'POST':
        code = request.POST['account']
        family = request.POST['family']
        classe = request.POST['classe']
        Label = request.POST['Label']
        OHADA_Code = request.POST['OHADA_Code']
        note = request.POST['note']
        if account.account != code:
            a = Account.objects.filter(account=code)
            if a.exists():
                context['error_message'] = 'Ce Compte OHADA existe deja.'
                return render(request, 'account/change_account.html', context)
        account.account = code
        account.family = family
        account.classe = classe
        account.Label = Label
        account.OHADA_Code = OHADA_Code
        account.note = note
        account.save()
        
        return redirect('view_account')
    return render(request, 'account/change_account.html', context)

@permission_required('nassib.delete_account', raise_exception=True)
def delete_account(request, account_id):
    account = get_object_or_404(Account, id=account_id)
    account.delete()
    return redirect('view_account')

@permission_required('nassib.view_agency', raise_exception=True)
def view_agency(request):
    search_query = request.POST.get('search', '')
    agency = Agency.objects.all()
    if search_query:  
        agency = agency.filter(Q(agency__icontains=search_query) | Q(authorisation__icontains=search_query))

    agency = agency.order_by('agency')
    paginator = Paginator(agency, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'agencies': page_obj,
        'search_query':search_query
    }
    return render(request, 'agency/view_agency.html', context)

@permission_required('nassib.add_agency', raise_exception=True)
def add_agency(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current
    }
    if request.method == 'POST':
        name = request.POST['agency']
        authorisation = request.POST['authorisation']
        product = Agency.objects.filter(agency=name)
        if product.exists():
            context['error_message'] = 'Cet agence existe deja.'
            return render(request, 'agency/add_agency.html', context)
        
        agency = Agency.objects.create(
            agency=name,
            authorisation=authorisation
        )
        agency.save()
        return redirect('view_agency')
    return render(request, 'agency/add_agency.html', context)

@permission_required('nassib.add_agency', raise_exception=True)
def import_agency(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current
    }

    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            context['error_message'] = 'Veuillez sélectionner un fichier à importer.'
            return render(request, 'agency/import_agency.html', context)
        
        excel_file = request.FILES['excel_file']
        try:
            workbook = load_workbook(excel_file)
            sheet = workbook.active
            
            rows = iter(sheet.rows)
            next(rows)  # Skip the header row
            for row in rows:
                name = row[0].value
                authorisation = row[1].value
                a = Agency.objects.filter(agency=name)
                if a.exists():
                    context['error_message'] = f'L\'agence {name} existe déjà.'
                    return render(request, 'agency/import_agency.html', context)
                agency = Agency.objects.create(
                    agency=name,
                    authorisation=authorisation
                )
                agency.save()
            return redirect('view_agency')
        except Exception as e:
            context['error_message'] = f'Une erreur s\'est produite lors de l\'importation des données : {str(e)}'
            return render(request, 'agency/import_agency.html', context)
        
    return render(request, 'agency/import_agency.html', context)

@permission_required('nassib.change_agency', raise_exception=True)
def change_agency(request,agency_id):
    agency = get_object_or_404(Agency, id=agency_id)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'agency':agency
    }
    if request.method == 'POST':
        name = request.POST['agency']
        authorisation = request.POST['authorisation']
        if agency.agency != name:
            a = Agency.objects.filter(agency=name)
            if a.exists():
                context['error_message'] = 'Cet agence existe deja.'
                return render(request, 'agency/change_agency.html', context)
        agency.agency = name
        agency.authorisation = authorisation
        agency.save()
        
        return redirect('view_agency')
    return render(request, 'agency/change_agency.html', context)

@permission_required('nassib.delete_agency', raise_exception=True)
def delete_agency(request, agency_id):
    agency = get_object_or_404(Agency, id=agency_id)
    agency.delete()
    return redirect('view_agency')

@permission_required('nassib.view_production', raise_exception=True)
def view_production(request):
    search_query = request.POST.get('search', '')
    production = Production.objects.all()
    if search_query:  
        production = production.filter(Q(ingredient__icontains=search_query) | Q(product__icontains=search_query) | Q(batch_no__icontains=search_query) | Q(product_type__icontains=search_query))

    production = production.order_by('-id')
    paginator = Paginator(production, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'productions': page_obj,
        'search_query':search_query
    }
    return render(request, 'production/view_production.html', context)

@permission_required('nassib.add_production', raise_exception=True)
def add_production(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    agencies = Agency.objects.all().order_by('agency')
    accounts = Account.objects.all().order_by('account')
    products = Product.objects.all().order_by('item')
    context = {
        'username': username_current,
        'agencies':agencies,
        'accounts':accounts,
        'products':products
    }
    if request.method == 'POST':
        date = request.POST['date']
        agency = request.POST['agency']
        product_type = request.POST['product_type']
        batch_no = request.POST['batch_no']
        product = request.POST['product']
        ingredient = request.POST['ingredient']
        quantity = parse_decimal(request.POST['quantity'])
        amount = parse_decimal(request.POST['amount'])
        journal = request.POST['journal']
        debit_account = request.POST['debit_account']
        debit_amount = parse_decimal(request.POST['debit_amount'])
        credit_account = request.POST['credit_account']
        credit_amount = parse_decimal(request.POST['credit_amount'])
        stock_out = parse_decimal(request.POST['stock_out'])
        if journal != "PRODUCTION":
            context['error_message'] = "le journal doit etre l'option 'PRODUCTION'"
            return render(request, 'production/add_production.html', context)
        
        production = Production.objects.create(
            date=date,
            agency=agency,
            product_type=product_type,
            batch_no=batch_no,
            product=product,
            ingredient=ingredient,
            quantity=quantity,
            amount=amount,
            journal=journal,
            debit_account=debit_account,
            debit_amount=debit_amount,
            credit_account=credit_account,
            credit_amount=credit_amount,
            stock_out=stock_out
        )
        production.save()
        try:
            stockagency = StockAgency.objects.get(ingredient=ingredient, agency=agency)
            stockagency.stock_out += stock_out
            stockagency.quantity = stockagency.stock_in - stockagency.stock_out
            stockagency.save()
        except StockAgency.DoesNotExist:
            context['error_message'] = "Une erreur s\'est produite lors de la recherche des données de cette ingredient."
            return render(request, 'production/add_production.html', context)
        try:
            stockglobal = StockGlobal.objects.get(ingredient=ingredient)
            stockglobal.stock_out += stock_out
            stockglobal.quantity = stockglobal.stock_in - stockglobal.stock_out
            stockglobal.amount = stockglobal.quantity * stockglobal.avg_price
            stockglobal.save()
        except StockGlobal.DoesNotExist:
            context['error_message'] = "Une erreur s\'est produite lors de la recherche des données de cette ingredient."
            return render(request, 'ventemarchandises/add_production.html', context)
        return redirect('view_production')
    return render(request, 'production/add_production.html', context)

@permission_required('nassib.change_production', raise_exception=True)
def change_production(request,production_id):
    production = get_object_or_404(Production, id=production_id)
    old_agency = production.agency
    old_ingredient = production.ingredient
    old_stock_out = parse_decimal(str(production.stock_out))
    def format_decimal(value):
        return str(value).replace(',', '.')
    production.quantity = format_decimal(production.quantity)
    
    current_user = request.user
    username_current = current_user.username.capitalize()
    agencies = Agency.objects.all().order_by('agency')
    accounts = Account.objects.all().order_by('account')
    products = Product.objects.all().order_by('item')
    context = {
        'username': username_current,
        'agencies':agencies,
        'accounts':accounts,
        'products':products,
        'production':production
    }
    if request.method == 'POST':
        date = request.POST['date']
        agency = request.POST['agency']
        product_type = request.POST['product_type']
        batch_no = request.POST['batch_no']
        product = request.POST['product']
        ingredient = request.POST['ingredient']
        quantity = parse_decimal(request.POST['quantity'])
        amount = parse_decimal(request.POST['amount'])
        journal = request.POST['journal']
        debit_account = request.POST['debit_account']
        debit_amount = parse_decimal(request.POST['debit_amount'])
        credit_account = request.POST['credit_account']
        credit_amount = parse_decimal(request.POST['credit_amount'])
        stock_out = parse_decimal(request.POST['stock_out'])
        if journal != "PRODUCTION":
            context['error_message'] = "le journal doit etre l'option 'PRODUCTION'"
            return render(request, 'production/change_production.html', context)
        
        production.date=date
        production.agency=agency
        production.product_type=product_type
        production.batch_no=batch_no
        production.product=product
        production.ingredient=ingredient
        production.quantity=quantity
        production.amount=amount
        production.journal=journal
        production.debit_account=debit_account
        production.debit_amount=debit_amount
        production.credit_account=credit_account
        production.credit_amount=credit_amount
        production.stock_out=stock_out
        production.save()
        if old_agency == agency and old_ingredient == ingredient:
            try:
                stockagency = StockAgency.objects.get(ingredient=ingredient, agency=agency)
                stockagency.stock_out += stock_out - old_stock_out
                stockagency.quantity = stockagency.stock_in - stockagency.stock_out
                stockagency.save()
            except StockAgency.DoesNotExist:
                context['error_message'] = "Une erreur s\'est produite lors de la recherche des données de cette ingredient."
                return render(request, 'production/add_production.html', context)
            try:
                stockglobal = StockGlobal.objects.get(ingredient=ingredient)
                stockglobal.stock_out += stock_out - old_stock_out
                stockglobal.quantity = stockglobal.stock_in - stockglobal.stock_out
                stockglobal.amount = stockglobal.quantity * stockglobal.avg_price
                stockglobal.save()
            except StockGlobal.DoesNotExist:
                context['error_message'] = "Une erreur s\'est produite lors de la recherche des données de cette ingredient."
                return render(request, 'ventemarchandises/add_production.html', context)
        else:
            try:
                old_stockagency = StockAgency.objects.get(ingredient=old_ingredient, agency=old_agency)
                old_stockagency.stock_out -= old_stock_out
                old_stockagency.quantity = old_stockagency.stock_in - old_stockagency.stock_out
                old_stockagency.save()
            except StockAgency.DoesNotExist:
                context['error_message'] = "Une erreur s\'est produite lors de la recherche des données de cette ingredient."
                return render(request, 'production/add_production.html', context)
            try:
                old_stockglobal = StockGlobal.objects.get(ingredient=old_ingredient)
                old_stockglobal.stock_out -= old_stock_out
                old_stockglobal.quantity = old_stockglobal.stock_in - old_stockglobal.stock_out
                old_stockglobal.amount = old_stockglobal.quantity * old_stockglobal.avg_price
                old_stockglobal.save()
            except StockGlobal.DoesNotExist:
                context['error_message'] = "Une erreur s\'est produite lors de la recherche des données de cette ingredient."
                return render(request, 'ventemarchandises/add_production.html', context)
            try:
                stockagency = StockAgency.objects.get(ingredient=ingredient, agency=agency)
                stockagency.stock_out += stock_out
                stockagency.quantity = stockagency.stock_in - stockagency.stock_out
                stockagency.save()
            except StockAgency.DoesNotExist:
                context['error_message'] = "Une erreur s\'est produite lors de la recherche des données de cette ingredient."
                return render(request, 'production/add_production.html', context)
            try:
                stockglobal = StockGlobal.objects.get(ingredient=ingredient)
                stockglobal.stock_out += stock_out
                stockglobal.quantity = stockglobal.stock_in - stockglobal.stock_out
                stockglobal.amount = stockglobal.quantity * stockglobal.avg_price
                stockglobal.save()
            except StockGlobal.DoesNotExist:
                context['error_message'] = "Une erreur s\'est produite lors de la recherche des données de cette ingredient."
                return render(request, 'ventemarchandises/add_production.html', context)
        return redirect('view_production')
    return render(request, 'production/change_production.html', context)

@permission_required('nassib.delete_production', raise_exception=True)
def delete_production(request, production_id):
    production = get_object_or_404(Production, id=production_id)
    if production.date == date.today():
        stockagency = StockAgency.objects.get(ingredient=production.ingredient, agency=production.agency)
        stockagency.stock_out -= production.stock_out
        stockagency.quantity = stockagency.stock_in - stockagency.stock_out
        stockagency.save()
        stockglobal = StockGlobal.objects.get(ingredient=production.ingredient)
        stockglobal.stock_out -= production.stock_out
        stockglobal.quantity = stockglobal.stock_in - stockglobal.stock_out
        stockglobal.amount = stockglobal.quantity * stockglobal.avg_price
        stockglobal.save()
    production.delete()
    return redirect('view_production')

@permission_required('nassib.view_batchproduction', raise_exception=True)
def view_batchproduction(request):
    search_query = request.POST.get('search', '')
    batchproduction = BatchProduction.objects.all()
    if search_query:  
        batchproduction = batchproduction.filter(Q(ingredient__icontains=search_query) | Q(product__icontains=search_query) | Q(batch_no__icontains=search_query))

    batchproduction = batchproduction.order_by('-id')
    paginator = Paginator(batchproduction, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'batchproductions': page_obj,
        'search_query':search_query
    }
    return render(request, 'batchproduction/view_batchproduction.html', context)

@permission_required('nassib.add_batchproduction', raise_exception=True)
def add_batchproduction(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    accounts = Account.objects.all().order_by('account')
    context = {
        'username': username_current,
        'accounts':accounts,
    }
    if request.method == 'POST':
        date = request.POST['date']
        agency = request.POST['agency']
        batch_no = request.POST['batch_no']
        product = request.POST['product']
        journal = request.POST['journal']
        debit_account = request.POST['debit_account']
        debit_amount = parse_decimal(request.POST['debit_amount'])
        credit_account = request.POST['credit_account']
        credit_amount = parse_decimal(request.POST['credit_amount'])
        total_amount = parse_decimal(request.POST['total_amount'])
        total_production =parse_decimal(request.POST['total_production'])
        unit_material_cost = parse_decimal(request.POST['unit_material_cost'])
        fixed_cost = parse_decimal(request.POST['fixed_cost'])
        variable_cost = parse_decimal(request.POST['variable_cost'])
        total_cost = parse_decimal(request.POST['total_cost'])
        total_fixed_cost = parse_decimal(request.POST['total_fixed_cost'])
        total_variable_cost = parse_decimal(request.POST['total_variable_cost'])
        batchproduction = BatchProduction.objects.filter(batch_no=batch_no,date=date,product=product,agency=agency)
        if batchproduction.exists():
            context['error_message'] = "cette production par lot a été deja effectuer."
            return render(request, 'batchproduction/add_batchproduction.html', context)
        batchproduction = BatchProduction.objects.create(
            date = date,
            agency = agency,
            batch_no = batch_no,
            product = product,
            journal = journal,
            debit_account = debit_account,
            debit_amount = debit_amount,
            credit_account = credit_account,
            credit_amount = credit_amount,
            total_amount = total_amount,
            total_production = total_production,
            unit_material_cost = unit_material_cost,
            fixed_cost = fixed_cost,
            variable_cost = variable_cost,
            total_cost = total_cost,
            total_fixed_cost = total_fixed_cost,
            total_variable_cost = total_variable_cost
        )
        batchproduction.save()
        product_type = Production.objects.filter(batch_no=batch_no).values_list('product_type', flat=True).distinct().first()
        if product_type == "Semi-finis":
            try:
                stockglobal = StockGlobal.objects.get(ingredient=product)
                stockglobal.amount += total_amount
                stockglobal.quantity += total_production
                stockglobal.stock_in += total_production
                stockglobal.avg_price = stockglobal.amount / stockglobal.quantity 
            except StockGlobal.DoesNotExist:
                stockglobal = StockGlobal(
                    ingredient=product,
                    amount=total_amount,
                    quantity=total_production,
                    stock_in=total_production,
                    avg_price=unit_material_cost,
                    stock_out=0
                )
            stockglobal.save()

            try:
                stockagency = StockAgency.objects.get(ingredient=product, agency=agency)
                stockagency.quantity += total_production
                stockagency.stock_in += total_production
                stockagency.last_update_date = timezone.now()
            except StockAgency.DoesNotExist:
                stockagency = StockAgency(
                    agency=agency,
                    ingredient=product,
                    quantity=total_production,
                    stock_in=total_production,
                    stock_out=0,
                    last_update_date=timezone.now()
                )
            stockagency.save()
        
        return redirect('view_batchproduction')
    return render(request, 'batchproduction/add_batchproduction.html', context)

@permission_required('nassib.delete_batchproduction', raise_exception=True)
def delete_batchproduction(request, batchproduction_id):
    batchproduction = get_object_or_404(BatchProduction, id=batchproduction_id)
    if batchproduction.date == date.today():
        product_type = Production.objects.filter(batch_no=batchproduction.batch_no).values_list('product_type', flat=True).distinct().first()
        if product_type == "Semi-finis":
            stockagency = StockAgency.objects.get(ingredient=batchproduction.product, agency=batchproduction.agency)
            stockagency.stock_in -= batchproduction.total_production
            stockagency.quantity = stockagency.stock_in - stockagency.stock_out
            stockagency.save()
            stockglobal = StockGlobal.objects.get(ingredient=batchproduction.product)
            stockglobal.stock_in -= batchproduction.total_production
            stockglobal.quantity = stockglobal.stock_in - stockglobal.stock_out
            stockglobal.amount = stockglobal.quantity * stockglobal.avg_price
            stockglobal.save()
    batchproduction.delete()
    return redirect('view_batchproduction')

@login_required
def get_batch_no(request, date):
    try:
        production = Production.objects.filter(date=date).values('batch_no').distinct()
        batch_numbers = [prod['batch_no'] for prod in production]
    except Production.DoesNotExist:
        batch_numbers = []
    except ValueError:
        batch_numbers = []

    return JsonResponse({'batch_numbers': batch_numbers})

@login_required
def get_batchno_agency(request, date,batch_no):
    try:
        production = Production.objects.filter(date=date,batch_no=batch_no).values('agency').distinct()
        agencies = [prod['agency'] for prod in production]
    except Production.DoesNotExist:
        agencies = []
    except ValueError:
        agencies = []

    return JsonResponse({'agencies': agencies})

@login_required
def get_batchno_agency_product(request, date,batch_no,agency):
    try:
        production = Production.objects.filter(date=date,batch_no=batch_no,agency=agency).values('product').distinct()
        products = [prod['product'] for prod in production]
    except Production.DoesNotExist:
        products = []
    except ValueError:
        products = []

    return JsonResponse({'products': products})

@login_required
def get_total_amount(request, date, batch_no, agency, product):
    try:
        production = Production.objects.filter(date=date, batch_no=batch_no, agency=agency, product=product)
        total_amount = production.aggregate(Sum('amount'))['amount__sum'] or 0
    except Production.DoesNotExist:
        total_amount = 0
    except ValueError:
        total_amount = 0

    return JsonResponse({'total_amount': total_amount})

@permission_required('nassib.view_finishedproduction', raise_exception=True)
def view_finishedproduction(request):
    search_query = request.POST.get('search', '')
    finishedproduction = FinishedProduction.objects.all()
    if search_query:  
        finishedproduction = finishedproduction.filter(Q(agency__icontains=search_query) | Q(product__icontains=search_query) | Q(reference__icontains=search_query))

    finishedproduction = finishedproduction.order_by('-id')
    paginator = Paginator(finishedproduction, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'finishedproductions': page_obj,
        'search_query':search_query
    }
    return render(request, 'finishedproduction/view_finishedproduction.html', context)

@permission_required('nassib.add_finishedproduction', raise_exception=True)
def add_finishedproduction(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    accounts = Account.objects.all().order_by('account')
    context = {
        'username': username_current,
        'accounts':accounts,
    }
    if request.method == 'POST':
        date = request.POST['date']
        agency = request.POST['agency']
        reference = request.POST['reference']
        product = request.POST['product']
        selling_price = parse_decimal(request.POST['selling_price'])
        quantity = parse_decimal(request.POST['quantity'])
        total_amount = parse_decimal(request.POST['total_amount'])
        debit_account = request.POST['debit_account']
        debit_amount = parse_decimal(request.POST['debit_amount'])
        credit_account = request.POST['credit_account']
        credit_amount = parse_decimal(request.POST['credit_amount'])
        
        finishedproduction = FinishedProduction.objects.create(
            date=date,
            agency=agency,
            reference=reference,
            product=product,
            selling_price=selling_price,
            quantity=quantity,
            total_amount=total_amount,
            debit_account=debit_account,
            debit_amount=debit_amount,
            credit_account=credit_account,
            credit_amount=credit_amount,
        )
        finishedproduction.save()
        return redirect('view_finishedproduction')
    return render(request, 'finishedproduction/add_finishedproduction.html', context)

@login_required
def get_batchnoproduction_by_date(request, date):
    try:
        finished_agencies = Production.objects.filter(date=date, product_type="Finis").values_list('agency', flat=True).distinct()
        
        batchproduction = BatchProduction.objects.filter(date=date, agency__in=finished_agencies).values('agency').distinct()
        agencies = [prod['agency'] for prod in batchproduction]
    except Production.DoesNotExist:
        agencies = []
    except BatchProduction.DoesNotExist:
        agencies = []
    except ValueError:
        agencies = []

    return JsonResponse({'agencies': agencies})

@login_required
def get_batchnoproduction_by_date_agency(request, date, agency):
    try:
        finished_products = Production.objects.filter(date=date, agency=agency, product_type="Finis").values_list('product', flat=True).distinct()
        
        batchproduction = BatchProduction.objects.filter(date=date, agency=agency, product__in=finished_products).values('product').distinct()
        products = [prod['product'] for prod in batchproduction]
    except Production.DoesNotExist:
        products = []
    except BatchProduction.DoesNotExist:
        products = []
    except ValueError:
        products = []

    return JsonResponse({'products': products})

@login_required
def get_price_by_product(request,product):
    try:
        pro = Product.objects.get(item=product)
        selling_price = pro.price
    except Product.DoesNotExist:
        selling_price = ""
    except ValueError:
        selling_price = ""

    return JsonResponse({'selling_price': selling_price})

@permission_required('nassib.delete_finishedproduction', raise_exception=True)
def delete_finishedproduction(request, finishedproduction_id):
    finishedproduction = get_object_or_404(FinishedProduction, id=finishedproduction_id)
    finishedproduction.delete()
    return redirect('view_finishedproduction')

@permission_required('nassib.view_distribution', raise_exception=True)
def view_distribution(request):
    search_query = request.POST.get('search', '')
    distribution = Distribution.objects.all()
    if search_query:  
        distribution = distribution.filter(Q(batch_no__icontains=search_query) | Q(product__icontains=search_query) | Q(initiator__icontains=search_query))

    distribution = distribution.order_by('-id')
    paginator = Paginator(distribution, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'distributions': page_obj,
        'search_query':search_query
    }
    return render(request, 'distribution/view_distribution.html', context)

@permission_required('nassib.add_distribution', raise_exception=True)
def add_distribution(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    agencies = Agency.objects.all().order_by('agency')
    distributions = Distribution.objects.all()
    total_stock_out = distributions.aggregate(total=Sum('stock_out'))['total'] or 0
    context = {
        'username': username_current,
        'agencies':agencies,
        'total_stock_out':total_stock_out
    }
    if request.method == 'POST':
        date = request.POST['date']
        batch_no =  request.POST['batch_no']
        product =  request.POST['product']
        initiator =  request.POST['initiator']
        agency =  request.POST['agency']
        total_production =  request.POST['total_production']
        quantity_sold = parse_decimal(request.POST['quantity_sold'])
        selling_price = parse_decimal(request.POST['selling_price'])
        sale_value = parse_decimal(request.POST['sale_value'])
        stock_out = parse_decimal(request.POST['stock_out'])
        stock_in = parse_decimal(request.POST['stock_in'])
        total_remaining_to_be_sold = parse_decimal(request.POST['total_remaining_to_be_sold'])
        if total_remaining_to_be_sold < 0:
            context['error_message'] = "La quantité totale vendue ne peut pas dépasser la production totale."
            return render(request, 'distribution/add_distribution.html', context)
        distribution=Distribution.objects.create(
            date=date,
            batch_no=batch_no,
            product=product,
            initiator=initiator,
            agency=agency,
            total_production=total_production,
            quantity_sold=quantity_sold,
            selling_price=selling_price,
            sale_value=sale_value,
            stock_out=stock_out,
            stock_in=stock_in
        )
        distribution.save()
        return redirect('view_distribution')
    return render(request, 'distribution/add_distribution.html', context)

@login_required
def get_batch_no_production(request, date):
    try:
        finished_batches = Production.objects.filter(date=date, product_type="Finis").values_list('batch_no', flat=True).distinct()
        
        batchproduction = BatchProduction.objects.filter(date=date, batch_no__in=finished_batches).values('batch_no').distinct()
        batch_numbers = [prod['batch_no'] for prod in batchproduction]
    except Production.DoesNotExist:
        batch_numbers = []
    except BatchProduction.DoesNotExist:
        batch_numbers = []
    except ValueError:
        batch_numbers = []

    return JsonResponse({'batch_numbers': batch_numbers})

@login_required
def get_batchno_initiator(request, date,batch_no):
    try:
        batchproduction = BatchProduction.objects.filter(date=date,batch_no=batch_no).values('agency').distinct()
        initiators = [prod['agency'] for prod in batchproduction]
    except BatchProduction.DoesNotExist:
        initiators = []
    except ValueError:
        initiators = []

    return JsonResponse({'initiators': initiators})

@login_required
def get_batchno_initiator_product(request, date,batch_no,initiator):
    try:
        batchproduction = BatchProduction.objects.filter(date=date,batch_no=batch_no,agency=initiator).values('product').distinct()
        products = [prod['product'] for prod in batchproduction]
    except BatchProduction.DoesNotExist:
        products = []
    except ValueError:
        products = []

    return JsonResponse({'products': products})

@login_required
def get_total_production_price(request, batch_no, date, initiator, product):
    try:
        p = Product.objects.get(item=product)
        
        batch_production = BatchProduction.objects.get(batch_no=batch_no,date=date,agency=initiator,product=product)
        
        price = p.price
        total_production = batch_production.total_production
    except Product.DoesNotExist:
        price = 0
        total_production = 0
    except BatchProduction.DoesNotExist:
        price = 0
        total_production = 0
    except ValueError:
        price = 0
        total_production = 0

    return JsonResponse({'price': price, 'total_production': total_production})

@permission_required('nassib.delete_distribution', raise_exception=True)
def delete_distribution(request, distribution_id):
    distribution = get_object_or_404(Distribution, id=distribution_id)
    distribution.delete()
    return redirect('view_distribution')

@permission_required('nassib.view_salescontrol', raise_exception=True)
def view_salescontrol(request):
    search_query = request.POST.get('search', '')
    salescontrols = SalesControl.objects.all()
    if search_query:  
        salescontrols = salescontrols.filter(Q(date__icontains=search_query) | Q(agency__icontains=search_query))

    salescontrols = salescontrols.order_by('-id')
    paginator = Paginator(salescontrols, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'salescontrols': page_obj,
        'search_query':search_query
    }
    return render(request, 'salescontrol/view_salescontrol.html', context)

@permission_required('nassib.add_salescontrol', raise_exception=True)
def add_salescontrol(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
    }
    if request.method == 'POST':
        date =  request.POST['date']
        agency = request.POST['agency']
        opening_balance = parse_decimal(request.POST['opening_balance'])
        sales_made= parse_decimal(request.POST['sales_made'])
        total_due = parse_decimal(request.POST['total_due'])
        cash_collection = parse_decimal(request.POST['cash_collection'])
        credit_sales = parse_decimal(request.POST['credit_sales'])
        donation = parse_decimal(request.POST['donation'])
        domaged_goods = parse_decimal(request.POST['domaged_goods'])
        closing_stock = parse_decimal(request.POST['closing_stock'])
        total_collection = parse_decimal(request.POST['total_collection'])
        collection_difference = parse_decimal(request.POST['collection_difference'])
        salescontrol = SalesControl.objects.filter(date=date,agency=agency)
        if salescontrol.exists():
            context['error_message'] = "Ce contrôle de vente a déjà été effectué."
            return render(request, 'salescontrol/add_salescontrol.html', context)
        salescontrol = SalesControl.objects.create(
            date=date,
            agency=agency,
            opening_balance=opening_balance,
            sales_made=sales_made,
            total_due=total_due,
            cash_collection=cash_collection,
            credit_sales=credit_sales,
            donation=donation,
            domaged_goods=domaged_goods,
            closing_stock=closing_stock,
            total_collection=total_collection,
            collection_difference=collection_difference,
        )
        salescontrol.save()
        stocfp = StockFP.objects.create(
            agency = agency,
            date = date,
            closing_stock = closing_stock
        )
        stocfp.save()
        return redirect('view_salescontrol')
    return render(request, 'salescontrol/add_salescontrol.html', context)

@login_required
def get_date_agency(request, date):
    try:
        distribution = Distribution.objects.filter(date=date).values('agency').distinct()
        agencies = [prod['agency'] for prod in distribution]
    except Distribution.DoesNotExist:
        agencies = []
    except ValueError:
        agencies = []

    return JsonResponse({'agencies': agencies})

@login_required
def get_date_agency_sales_made(request, date, agency):
    try:
        comptafp = ComptaFp.objects.filter(date=date, agency=agency)
        
        sales_made = comptafp.aggregate(total_sales=Sum('sale_value'))['total_sales'] or 0
    except ValueError:
        sales_made = 0

    return JsonResponse({'sales_made': sales_made})

@login_required
def get_date_agency_opening_balance(request,agency):
    try:
        stockfp = StockFP.objects.get(agency=agency)
        opening_balance=stockfp.closing_stock
    except StockFP.DoesNotExist:
        opening_balance = 0
    except ValueError:
        opening_balance = 0

    return JsonResponse({'opening_balance': opening_balance})

@permission_required('nassib.delete_salescontrol', raise_exception=True)
def delete_salescontrol(request, salescontrol_id):
    salescontrol = get_object_or_404(SalesControl, id=salescontrol_id)
    if salescontrol.date == date.today():
        stocfp = get_object_or_404(StockFP, agency=salescontrol.agency)
        stocfp.closing_stock = salescontrol.opening_balance
        stocfp.date = date.today() - timedelta(days=1)
        if stocfp.closing_stock == parse_decimal('0'):
            stocfp.delete()
        else:
            stocfp.save()
    salescontrol.delete()
    return redirect('view_salescontrol')

@permission_required('nassib.view_comptafp', raise_exception=True)
def view_comptafp(request):
    search_query = request.POST.get('search', '')
    comptafps = ComptaFp.objects.all()
    if search_query:  
        comptafps = comptafps.filter(Q(date__icontains=search_query) | Q(agency__icontains=search_query))

    comptafps = comptafps.order_by('-id')
    paginator = Paginator(comptafps, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'comptafps': page_obj,
        'search_query':search_query
    }
    return render(request, 'comptafp/view_comptafp.html', context)

@permission_required('nassib.add_comptafp', raise_exception=True)
def add_comptafp(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    accounts = Account.objects.all().order_by('account')
    context = {
        'username': username_current,
        'accounts':accounts
    }
    if request.method == 'POST':
        date =  request.POST['date']
        agency = request.POST['agency']
        journal = request.POST['journal']
        sale_value = parse_decimal(request.POST['sale_value'])
        debit_account =  request.POST['debit_account']
        credit_account =  request.POST['credit_account']
        debit_amount= parse_decimal(request.POST['debit_amount'])
        credit_amount= parse_decimal(request.POST['credit_amount'])
        comptafp = ComptaFp.objects.filter(date=date,agency=agency)
        if comptafp.exists():
            context['error_message'] = "Cette comptabilité pour ce produit fini a déjà été effectuée."
            return render(request, 'comptafp/add_comptafp.html', context)
        comptafp = ComptaFp.objects.create(
            date=date,
            agency=agency,
            journal=journal,
            sale_value=sale_value,
            debit_account=debit_account,
            credit_account=credit_account,
            debit_amount=debit_amount,
            credit_amount=credit_amount
        )
        comptafp.save()
        return redirect('view_comptafp')
    return render(request, 'comptafp/add_comptafp.html', context)

@login_required
def get_date_agency_sale_value(request, date,agency):
    try:
        distribution = Distribution.objects.get(date=date,agency=agency)
        sale_value=distribution.sale_value
    except Distribution.DoesNotExist:
        sale_value = 0
    except ValueError:
        sale_value = 0

    return JsonResponse({'sale_value': sale_value})

@permission_required('nassib.delete_comptafp', raise_exception=True)
def delete_comptafp(request, comptafp_id):
    comptafp = get_object_or_404(ComptaFp, id=comptafp_id)
    comptafp.delete()
    return redirect('view_comptafp')

@permission_required('nassib.view_comptapos', raise_exception=True)
def view_comptapos(request):
    search_query = request.POST.get('search', '')
    comptapos = ComptaPos.objects.all()
    if search_query:  
        comptapos = comptapos.filter(Q(date__icontains=search_query) | Q(agency__icontains=search_query))

    comptapos = comptapos.order_by('-id')
    paginator = Paginator(comptapos, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'comptapos': page_obj,
        'search_query':search_query
    }
    return render(request, 'comptapos/view_comptapos.html', context)

@permission_required('nassib.add_comptapos', raise_exception=True)
def add_comptapos(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    accounts = Account.objects.all().order_by('account')
    context = {
        'username': username_current,
        'accounts':accounts
    }
    if request.method == 'POST':
        date =  request.POST['date']
        agency = request.POST['agency']
        cash_collection= parse_decimal(request.POST['cash_collection'])
        domaged_goods = parse_decimal(request.POST['domaged_goods'])
        donation = parse_decimal(request.POST['donation'])
        total = parse_decimal(request.POST['total'])
        debit_account =  request.POST['debit_account']
        credit_account =  request.POST['credit_account']
        debit_amount= parse_decimal(request.POST['debit_amount'])
        credit_amount= parse_decimal(request.POST['credit_amount'])
        comptapos = ComptaPos.objects.filter(date=date,agency=agency)
        if comptapos.exists():
            context['error_message'] = "Cette comptabilité pos a déjà été effectuée."
            return render(request, 'comptapos/add_comptapos.html', context)
        comptapos = ComptaPos.objects.create(
            date=date,
            agency=agency,
            cash_collection=cash_collection,
            domaged_goods=domaged_goods,
            donation=donation,
            total=total,
            debit_account=debit_account,
            credit_account=credit_account,
            debit_amount=debit_amount,
            credit_amount=credit_amount
        )
        comptapos.save()
        return redirect('view_comptapos')
    return render(request, 'comptapos/add_comptapos.html', context)

@login_required
def get_date_by_agency(request, date):
    try:
        salescontrol = SalesControl.objects.filter(date=date).values('agency').distinct()
        agencies = [prod['agency'] for prod in salescontrol]
    except SalesControl.DoesNotExist:
        agencies = []
    except ValueError:
        agencies = []

    return JsonResponse({'agencies': agencies})

@login_required
def get_compta_pos_afterdate(request, date,agency):
    try:
        salescontrol = SalesControl.objects.get(date=date,agency=agency)
        cash_collection = salescontrol.cash_collection
        domaged_goods = salescontrol.domaged_goods
        donation = salescontrol.donation
        total = cash_collection + donation +domaged_goods
    except SalesControl.DoesNotExist:
        cash_collection = 0
        domaged_goods = 0
        donation = 0
        total = 0
    except ValueError:
        cash_collection = 0
        domaged_goods = 0
        donation = 0
        total = 0

    return JsonResponse({'cash_collection': cash_collection,'domaged_goods':domaged_goods,'donation':donation,'total':total})

@permission_required('nassib.delete_comptapos', raise_exception=True)
def delete_comptapos(request, comptapos_id):
    comptapos = get_object_or_404(ComptaPos, id=comptapos_id)
    comptapos.delete()
    return redirect('view_comptapos')

@permission_required('nassib.view_cashbank', raise_exception=True)
def view_cashbank(request):
    search_query = request.POST.get('search', '')
    cashbanks = CashBank.objects.all()
    if search_query:  
        cashbanks = cashbanks.filter(Q(date__icontains=search_query) | Q(detail__icontains=search_query))

    cashbanks = cashbanks.order_by('-id')
    paginator = Paginator(cashbanks, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'cashbanks': page_obj,
        'search_query':search_query
    }
    return render(request, 'cashbank/view_cashbank.html', context)

@permission_required('nassib.add_cashbank', raise_exception=True)
def add_cashbank(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    accounts = Account.objects.all().order_by('account')
    context = {
        'username': username_current,
        'accounts':accounts
    }
    if request.method == 'POST':
        date =  request.POST['date']
        detail = request.POST['detail']
        journal = request.POST['journal']
        amount= parse_decimal(request.POST['amount'])
        debit_account =  request.POST['debit_account']
        credit_account =  request.POST['credit_account']
        debit_amount= parse_decimal(request.POST['debit_amount'])
        credit_amount= parse_decimal(request.POST['credit_amount'])
        cashbank = CashBank.objects.create(
            date=date,
            detail=detail,
            journal=journal,
            amount=amount,
            debit_account=debit_account,
            credit_account=credit_account,
            debit_amount=debit_amount,
            credit_amount=credit_amount
        )
        cashbank.save()
        return redirect('view_cashbank')
    return render(request, 'cashbank/add_cashbank.html', context)

@permission_required('nassib.delete_cashbank', raise_exception=True)
def delete_cashbank(request, cashbank_id):
    cashbank = get_object_or_404(CashBank, id=cashbank_id)
    cashbank.delete()
    return redirect('view_cashbank')

@permission_required('nassib.view_othersrenewed', raise_exception=True)
def view_othersrenewed(request):
    search_query = request.POST.get('search', '')
    othersrenewed = OthersRenewed.objects.all()
    if search_query:  
        othersrenewed = othersrenewed.filter(Q(date__icontains=search_query) | Q(detail__icontains=search_query))

    othersrenewed = othersrenewed.order_by('-id')
    paginator = Paginator(othersrenewed, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'othersreneweds': page_obj,
        'search_query':search_query
    }
    return render(request, 'othersrenewed/view_othersrenewed.html', context)

@permission_required('nassib.add_othersrenewed', raise_exception=True)
def add_othersrenewed(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    accounts = Account.objects.all().order_by('account')
    context = {
        'username': username_current,
        'accounts':accounts
    }
    if request.method == 'POST':
        date =  request.POST['date']
        detail = request.POST['detail']
        journal = request.POST['journal']
        amount= parse_decimal(request.POST['amount'])
        debit_account =  request.POST['debit_account']
        credit_account =  request.POST['credit_account']
        debit_amount= parse_decimal(request.POST['debit_amount'])
        credit_amount= parse_decimal(request.POST['credit_amount'])
        othersrenewed = OthersRenewed.objects.create(
            date=date,
            detail=detail,
            journal=journal,
            amount=amount,
            debit_account=debit_account,
            credit_account=credit_account,
            debit_amount=debit_amount,
            credit_amount=credit_amount
        )
        othersrenewed.save()
        return redirect('view_othersrenewed')
    return render(request, 'othersrenewed/add_othersrenewed.html', context)

@permission_required('nassib.delete_othersrenewed', raise_exception=True)
def delete_othersrenewed(request, othersrenewed_id):
    othersrenewed = get_object_or_404(OthersRenewed, id=othersrenewed_id)
    othersrenewed.delete()
    return redirect('view_othersrenewed')

@permission_required('nassib.view_comptavm', raise_exception=True)
def view_comptavm(request):
    search_query = request.POST.get('search', '')
    comptavm = ComptaVm.objects.all()
    if search_query:  
        comptavm = comptavm.filter(Q(agency__icontains=search_query) | Q(date__icontains=search_query) | Q(customer__icontains=search_query))

    comptavm = comptavm.order_by('-id')
    paginator = Paginator(comptavm, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'comptavms': page_obj,
        'search_query':search_query
    }
    return render(request, 'comptavm/view_comptavm.html', context)

@permission_required('nassib.add_comptavm', raise_exception=True)
def add_comptavm(request):
    current_user = request.user
    username_current = current_user.username.capitalize()
    accounts = Account.objects.all().order_by('account')
    context = {
        'username': username_current,
        'accounts':accounts
    }
    if request.method == 'POST':
        date =  request.POST['date']
        customer = request.POST['customer']
        agency = request.POST['agency']
        vante_marchandises= parse_decimal(request.POST['vante_marchandises'])
        debit_account =  request.POST['debit_account']
        credit_account =  request.POST['credit_account']
        debit_amount= parse_decimal(request.POST['debit_amount'])
        credit_amount= parse_decimal(request.POST['credit_amount'])
        comptavm = ComptaVm.objects.create(
            date=date,
            customer=customer,
            agency=agency,
            vante_marchandises=vante_marchandises,
            debit_account=debit_account,
            credit_account=credit_account,
            debit_amount=debit_amount,
            credit_amount=credit_amount
        )
        comptavm.save()
        return redirect('view_comptavm')
    return render(request, 'comptavm/add_comptavm.html', context)

@permission_required('nassib.delete_comptavm', raise_exception=True)
def delete_comptavm(request, comptavm_id):
    comptavm = get_object_or_404(ComptaVm, id=comptavm_id)
    comptavm.delete()
    return redirect('view_comptavm')

@login_required
def get_date_customer(request, date):
    try:
        ventemarchandises = VenteMarchandises.objects.filter(date=date).values('customer').distinct()
        customers = [prod['customer'] for prod in ventemarchandises]
    except VenteMarchandises.DoesNotExist:
        customers = []
    except ValueError:
        customers = []

    return JsonResponse({'customers': customers})

@login_required
def get_date_customer_agency(request, date,customer):
    try:
        ventemarchandises = VenteMarchandises.objects.filter(date=date,customer=customer).values('agency').distinct()
        agencies = [prod['agency'] for prod in ventemarchandises]
    except VenteMarchandises.DoesNotExist:
        agencies = []
    except ValueError:
        agencies = []

    return JsonResponse({'agencies': agencies})

@login_required
def get_date_customer_agency_vante_marchandises(request, date,customer,agency):
    try:
        ventemarchandises = VenteMarchandises.objects.get(date=date,customer=customer,agency=agency)
        vante_marchandises = ventemarchandises.vante_marchandises
    except VenteMarchandises.DoesNotExist:
        vante_marchandises = 0
    except ValueError:
        vante_marchandises = 0

    return JsonResponse({'vante_marchandises': vante_marchandises})

@permission_required('nassib.view_consolidation', raise_exception=True)
def view_consolidation(request):
    search_query = request.POST.get('search', '')
    consolidation = Consolidation.objects.all()
    if search_query:
        consolidation = consolidation.filter(
            Q(agency__icontains=search_query) |  # Filtrer par agence
            Q(date__icontains=search_query) |    # Filtrer par date
            Q(total_debit__icontains=search_query) |  # Filtrer par total debit
            Q(total_credit__icontains=search_query)   # Filtrer par total credit
        )

    consolidation = consolidation.order_by('-id')
    paginator = Paginator(consolidation, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'consolidations': page_obj,
        'search_query':search_query
    }
    return render(request, 'consolidation/view_consolidation.html', context)

@permission_required('nassib.view_consolidation', raise_exception=True)
def consolidation(request):
    # Consolidation des achats
    achats = Achat.objects.all()
    for achat in achats:
        if not Consolidation.objects.filter(
            date=achat.date,
            agency=achat.agency,
            debit_account=achat.debit_account,
            credit_account=achat.credit_account
        ).exists():
            Consolidation.objects.create(
                date=achat.date,
                agency=achat.agency,
                debit_account=achat.debit_account,
                debit_amount=achat.debit_amount,
                credit_account=achat.credit_account,
                credit_amount=achat.credit_amount,
                product=achat.product,
                quantity=achat.quantity,
                unit_price=achat.unit_price,
                total_amount=achat.total_amount,
                provider=achat.provider,
                payment_method=achat.payment_method,
                reference=achat.reference
            )

    # Consolidation des productions
    productions = Production.objects.all()
    for production in productions:
        if not Consolidation.objects.filter(
            date=production.date,
            agency=production.agency,
            debit_account=production.debit_account,
            credit_account=production.credit_account
        ).exists():
            Consolidation.objects.create(
                date=production.date,
                agency=production.agency,
                debit_account=production.debit_account,
                debit_amount=production.debit_amount,
                credit_account=production.credit_account,
                credit_amount=production.credit_amount,
                product=production.product,
                quantity=production.quantity,
                unit_price=production.unit_price,
                total_amount=production.total_amount,
                batch_no=production.batch_no,
                department=production.department
            )

    # Consolidation des ventes de marchandises
    ventes = VenteMarchandises.objects.all()
    for vente in ventes:
        if not Consolidation.objects.filter(
            date=vente.date,
            agency=vente.agency,
            debit_account=vente.debit_account,
            credit_account=vente.credit_account
        ).exists():
            Consolidation.objects.create(
                date=vente.date,
                agency=vente.agency,
                debit_account=vente.debit_account,
                debit_amount=vente.debit_amount,
                credit_account=vente.credit_account,
                credit_amount=vente.credit_amount,
                vante_marchandises=vente.vante_marchandises,
                selling_price=vente.selling_price,
                customer=vente.customer
            )

    # Consolidation des finished productions
    finished_productions = FinishedProduction.objects.all()
    for finished in finished_productions:
        if not Consolidation.objects.filter(
            date=finished.date,
            agency=finished.agency,
            debit_account=finished.debit_account,
            credit_account=finished.credit_account
        ).exists():
            Consolidation.objects.create(
                date=finished.date,
                agency=finished.agency,
                debit_account=finished.debit_account,
                debit_amount=finished.debit_amount,
                credit_account=finished.credit_account,
                credit_amount=finished.credit_amount,
                product=finished.product,
                batch_no=finished.batch_no,
                quantity=finished.quantity,
                department=finished.department
            )

    # Consolidation des transferts
    transfers = Transfer.objects.all()
    for transfer in transfers:
        if not Consolidation.objects.filter(
            date=transfer.date,
            agency=transfer.agency,
            debit_account=transfer.debit_account,
            credit_account=transfer.credit_account
        ).exists():
            Consolidation.objects.create(
                date=transfer.date,
                agency=transfer.agency,
                debit_account=transfer.debit_account,
                debit_amount=transfer.debit_amount,
                credit_account=transfer.credit_account,
                credit_amount=transfer.credit_amount,
                detail=transfer.detail
            )

    # Consolidation de CashBank
    cash_bank = CashBank.objects.all()
    for cash in cash_bank:
        if not Consolidation.objects.filter(
            date=cash.date,
            agency=cash.agency,
            debit_account=cash.debit_account,
            credit_account=cash.credit_account
        ).exists():
            Consolidation.objects.create(
                date=cash.date,
                agency=cash.agency,
                debit_account=cash.debit_account,
                debit_amount=cash.debit_amount,
                credit_account=cash.credit_account,
                credit_amount=cash.credit_amount,
                opening_balance=cash.opening_balance,
                total_due=cash.total_due,
                cash_collection=cash.cash_collection
            )

    # Consolidation de SalesControl
    sales_control = SalesControl.objects.all()
    for sales in sales_control:
        if not Consolidation.objects.filter(
            date=sales.date,
            agency=sales.agency,
            debit_account=sales.debit_account,
            credit_account=sales.credit_account
        ).exists():
            Consolidation.objects.create(
                date=sales.date,
                agency=sales.agency,
                debit_account=sales.debit_account,
                debit_amount=sales.debit_amount,
                credit_account=sales.credit_account,
                credit_amount=sales.credit_amount,
                cost_of_sales=sales.cost_of_sales,
                total_collection=sales.total_collection,
                closing_stock=sales.closing_stock,
                customer=sales.customer
            )

    # Une fois la consolidation faite, rediriger vers la vue de consolidation
    return redirect('view_consolidation')

@permission_required('nassib.view_trialbalance', raise_exception=True)
def view_trialbalance(request):
    search_query = request.POST.get('search', '')
    trialbalance = TrialBalance.objects.all()
    if search_query:  
        trialbalance = trialbalance.filter(Q(customer__icontains=search_query) | Q(provider__icontains=search_query) | Q(product__icontains=search_query) | Q(agency__icontains=search_query) | Q(credit_account__icontains=search_query) | Q(debit_account__icontains=search_query))

    trialbalance = trialbalance.order_by('-id')
    paginator = Paginator(trialbalance, 12) 
    page_number = request.GET.get('page', 1) 
    page_obj = paginator.get_page(page_number)
    current_user = request.user
    username_current = current_user.username.capitalize()
    context = {
        'username': username_current,
        'trialbalances': page_obj,
        'search_query':search_query
    }
    return render(request, 'trialbalance/view_trialbalance.html', context)

@permission_required('nassib.view_trialbalance', raise_exception=True)
def trialbalance(request):
    # Initialiser un dictionnaire pour stocker les totaux par date, agence et comptes
    trialbalance_data = {}

    # Fonction pour ajouter les débits et crédits au dictionnaire
    def add_to_trialbalance_data(date, agency, debit_account, credit_account, debit, credit):
        key = (date, agency, debit_account, credit_account)
        if key not in trialbalance_data:
            trialbalance_data[key] = {'total_debit': 0, 'total_credit': 0}
        trialbalance_data[key]['total_debit'] += debit
        trialbalance_data[key]['total_credit'] += credit

    # Consolidation des données du modèle Consolidation
    consolidations = Consolidation.objects.all()
    for consolidation in consolidations:
        add_to_trialbalance_data(
            consolidation.date,
            consolidation.agency,
            consolidation.debit_account,  # Compte de débit
            consolidation.credit_account,  # Compte de crédit
            consolidation.debit_amount or 0,  # Gérer les valeurs nulles
            consolidation.credit_amount or 0  # Gérer les valeurs nulles
        )

    # Insérer ou mettre à jour les données dans TrialBalance
    for (date, agency, debit_account, credit_account), totals in trialbalance_data.items():
        total_debit = totals['total_debit']
        total_credit = totals['total_credit']
        balance = total_debit - total_credit

        # Vérifier si une entrée existe déjà pour cette date, agence, et comptes
        trial_balance_entry, created = TrialBalance.objects.get_or_create(
            date=date,
            agency=agency,
            debit_account=debit_account,  # Compte de débit
            credit_account=credit_account,  # Compte de crédit
            defaults={
                'debit_amount': total_debit,
                'credit_amount': total_credit,
            }
        )
        # Si l'entrée existe déjà, mettre à jour les champs
        if not created:
            trial_balance_entry.debit_amount += total_debit
            trial_balance_entry.credit_amount += total_credit
            trial_balance_entry.save()

    # Redirection après la consolidation
    return redirect('view_trialbalance')
